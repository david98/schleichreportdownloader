# coding=UTF-8

import serial
import time
import datetime
import logging

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from PyQt5 import QtCore, QtGui, QtWidgets


class NoReportException(Exception):
    pass


class TestReport:
    """ Represents a report

    name: Name of the preset
    date: Date and time of execution
    steps: A list of {
        'name': Name of the step
        'method': HV,
        'test_condition': Set voltage,
        'limit_value': Maximum current,
        'actual_condition': Actual measured voltage,
        'actual_value': Actual measured current,
        'test_duration': Step duration in seconds,
    }
    """

    def __init__(self, report_string):
        result = report_string.split('NUM_1')
        test_info = result[1].split(' ')
        self.name = test_info[1].replace('NAME_', '').replace('*', ' ')
        date_string = test_info[2].replace('DA_', '').replace('_', ' ')
        self.date = datetime.datetime.strptime(date_string, '%d.%m.%y %H:%M:%S')
        elements = result[0].split(' ')
        steps = [elements[n:n + 7] for n in range(0, len(elements), 7)]
        self.steps_with_results = []
        for step in steps:
            if len(step) > 1:
                del step[0]
                split_step_info = step[5].split('_')
                parsed_step = {
                    'name': split_step_info[2].replace('*', ' '),
                    'method': step[0],
                    'test_condition': float(step[1]),
                    'limit_value': float(step[2]),
                    'actual_condition': float(step[3]),
                    'actual_value': float(step[4]),
                    'test_duration': float(split_step_info[1]),
                }
                parsed_step['go'] = 'GO' if parsed_step['actual_value'] <= parsed_step['limit_value'] else 'NGO'
                self.steps_with_results.append(parsed_step)

    def store_as_xlsx(self, name):
        wb = Workbook()

        bold_font = Font(bold=True)

        dest_filename = name

        ws1 = wb.active
        ws1.title = "Test Report"

        ws1['A1'] = 'Preset Name'
        ws1['B1'] = 'Date'
        ws1['A1'].font = bold_font
        ws1['B1'].font = bold_font

        ws1['A2'] = self.name
        ws1['B2'] = self.date

        ws1['A4'] = 'Step Number'
        ws1['B4'] = 'Method'
        ws1['C4'] = 'Step Name'
        ws1['D4'] = 'Limit Value'
        ws1['E4'] = 'Actual Value'
        ws1['F4'] = 'Test Condition'
        ws1['G4'] = 'Actual Condition'
        ws1['H4'] = 'Test Duration'
        ws1['I4'] = 'Go'

        ws1['A4'].font = bold_font
        ws1['B4'].font = bold_font
        ws1['C4'].font = bold_font
        ws1['D4'].font = bold_font
        ws1['E4'].font = bold_font
        ws1['F4'].font = bold_font
        ws1['G4'].font = bold_font
        ws1['H4'].font = bold_font
        ws1['I4'].font = bold_font

        row = 5
        for step in self.steps_with_results:
            for col in range(1, len(step) + 2):
                value = 0
                if col == 1:
                    value = row - 4
                if col == 2:
                    value = step['method']
                elif col == 3:
                    value = step['name']
                elif col == 4:
                    value = str(step['limit_value']) + ' mA'
                elif col == 5:
                    value = str(step['actual_value']) + ' mA'
                elif col == 6:
                    value = str(step['test_condition']) + ' V'
                elif col == 7:
                    value = str(step['actual_condition']) + ' V'
                elif col == 8:
                    value = str(step['test_duration']) + ' s'
                elif col == 9:
                    value = step['go']
                _ = ws1.cell(column=col, row=row, value=value)
            row += 1

        column_widths = []
        for row in ws1.iter_rows():
            for i, cell in enumerate(row):
                try:
                    column_widths[i] = max(column_widths[i], len(as_text(cell.value)))
                except IndexError:
                    column_widths.append(len(as_text(cell.value)))

        for i, column_width in enumerate(column_widths):
            ws1.column_dimensions[get_column_letter(i + 1)].width = column_width + 5

        wb.save(filename=dest_filename)

    def __str__(self):
        string = ''
        string += self.name + ' | ' + str(self.date) + '\n'
        for idx, step in enumerate(self.steps_with_results):
            string += str(step) + '\n'
        return string


class TestingDevice:

    BEEP_COMMAND = [0x02, 0x81, 0xfa, 0x62, 0x20, 0x33, 0x42, 0x03]
    IDENTIFY_COMMAND = [0x02, 0x81, 0xfd]
    GET_REPORT_COMMAND = [0x02, 0x81, 0x06]
    START_TEST_COMMAND = [0x02, 0x81, 0xfa, 0x73, 0x20, 0x32, 0x41, 0x03]

    def __init__(self, serial_port: str):
        self.ser = serial.Serial(serial_port, baudrate=9600, timeout=None, parity=serial.PARITY_NONE,
                                 bytesize=serial.EIGHTBITS, stopbits=serial.STOPBITS_ONE, xonxoff=False)
        self.id_string = ""

    def send_custom_command(self, command_hex):
        self.ser.write(serial.to_bytes(command_hex))
        time.sleep(0.12)

    def read_all(self):
        read_data = ''
        while self.ser.in_waiting > 0:
            try:
                new_data = self.ser.read(self.ser.in_waiting)
                read_data += new_data.decode(errors='ignore')
            except TypeError:
                continue
            time.sleep(0.12)
        return read_data

    def beep(self):
        self.send_custom_command(TestingDevice.BEEP_COMMAND)

    def identify(self):
        self.send_custom_command(TestingDevice.IDENTIFY_COMMAND)
        id_string = self.read_all()
        id_string = (id_string.split("Conness.")[0])[3:]
        self.id_string = id_string
        return id_string

    def get_first_available_report(self):
        self.send_custom_command(TestingDevice.GET_REPORT_COMMAND)
        result = self.read_all()
        if len(result.strip().replace('\x07', '').replace('\x15', '').replace('4', '').replace('\x03', '').replace('2', '')) == 0:
            raise NoReportException('No report available for download.')
        return TestReport(result)

    def get_all_reports(self):
        reports = []
        while True:
            try:
                reports.append(self.get_first_available_report())
            except NoReportException:
                return reports

    def start_test(self):
        self.send_custom_command(TestingDevice.START_TEST_COMMAND)
        self.ser.reset_input_buffer()

    def clear_all_reports(self):
        self.get_all_reports()

    def close_communication(self):
        self.ser.close()


class TextFeedback:

    def __init__(self, qt_text_edit: QtWidgets.QTextEdit):
        self.qt_text_edit = qt_text_edit
        self._translate = QtCore.QCoreApplication.translate

    def set_text(self, text: str):
        self.qt_text_edit.setText(self._translate("MainWindow", text))

    def clear(self):
        self.set_text('')

    def append(self, text: str):
        current_text = self.qt_text_edit.toPlainText()
        self.set_text(current_text + text)
        pass

    def append_new_line(self, text: str):
        current_text = self.qt_text_edit.toPlainText()
        if len(current_text) == 0:
            self.append(text)
        else:
            self.append('\n' + text)


class StatusFeedback:

    def __init__(self, qt_label: QtWidgets.QLabel):
        self.qt_label = qt_label
        self._translate = QtCore.QCoreApplication.translate

    def set_text(self, text: str):
        self.qt_label.setText(self._translate("MainWindow", text))


class LoadingIndicator:

    def __init__(self, qt_label: QtWidgets.QLabel):
        self.qt_label = qt_label
        self.spinner = QtGui.QMovie('spinner.gif')
        self.visible = False

    def enable(self):
        if not self.visible:
            self.qt_label.setMovie(self.spinner)
            self.spinner.start()
            self.visible = True

    def disable(self):
        if self.visible:
            self.qt_label.clear()
            self.visible = False

    def toggle(self):
        if self.visible:
            self.disable()
        else:
            self.enable()


class StartTestControl:

    def __init__(self, qt_push_button: QtWidgets.QPushButton):
        self.qt_push_button = qt_push_button

    def enable(self):
        self.qt_push_button.setEnabled(True)

    def disable(self):
        self.qt_push_button.setEnabled(False)

    def toggle_enable(self):
        self.qt_push_button.setEnabled(not self.qt_push_button.isEnabled())


class TestManager:

    def __init__(self, device: TestingDevice):
        self.device = device
        self.text_feedback: TextFeedback = None
        self.status_feedback: StatusFeedback = None
        self.start_test_control: StartTestControl = None
        self.loading_indicator: LoadingIndicator = None

    def set_text_feedback(self, qt_text_edit: QtWidgets.QTextEdit):
        self.text_feedback = TextFeedback(qt_text_edit)

    def set_status_feedback(self, qt_label: QtWidgets.QLabel):
        self.status_feedback = StatusFeedback(qt_label)

    def set_start_test_control(self, qt_start_test_button: QtWidgets.QPushButton):
        self.start_test_control = StartTestControl(qt_start_test_button)

    def set_loading_indicator(self, qt_label: QtWidgets.QLabel):
        self.loading_indicator = LoadingIndicator(qt_label)

    def ready(self):
        self.status_feedback.set_text("Connected to {0}".format(self.device.identify()))

    def end_test(self):
        self.start_test_control.enable()
        self.loading_indicator.enable()

    def wait_for_report(self):
        report = None
        while True:
            QtCore.QCoreApplication.processEvents()
            try:
                report = self.device.get_first_available_report()
                break
            except NoReportException:
                continue
        self.text_feedback.append_new_line("Report downloaded succesfully.")
        report.store_as_xlsx("{0}.xlsx".format(int(time.time()*1000)))
        file_name = QtWidgets.QFileDialog.getSaveFileName(None, 'Save Report', './report-{0}.xlsx'.format(int(time.time()*1000)), filter='*.xlsx')
        report.store_as_xlsx(file_name)
        self.text_feedback.append_new_line("Report was saved to {0}".format(file_name))

    def start_test(self):
        self.start_test_control.disable()
        self.loading_indicator.disable()
        self.device.start_test()
        self.text_feedback.clear()
        self.text_feedback.append_new_line("Test started.")
        self.text_feedback.append_new_line("Waiting for report...")
        self.wait_for_report()
        return