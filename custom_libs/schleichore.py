# coding=UTF-8
from abc import ABC, abstractmethod

import datetime
import logging
import os
import time
from pathlib import Path

import serial
from PyQt5 import QtCore
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def as_text(value):
    if value is None:
        return ""
    return str(value)


class NoReportException(Exception):
    pass


class TestReport:
    """ Represents a report

    name: Name of the preset
    date: Date and time of execution
    steps: A list of {
        'name': Name of the step
        'method': HV,
        'test_condition': Set voltage,
        'limit_value': Maximum current,
        'actual_condition': Actual measured voltage,
        'actual_value': Actual measured current,
        'test_duration': Step duration in seconds,
    }
    """

    def __init__(self, report_string):
        result = report_string.split('NUM_1')
        test_info = result[1].split(' ')
        self.name = test_info[1].replace('NAME_', '').replace('*', ' ')
        date_string = test_info[2].replace('DA_', '').replace('_', ' ')
        self.date = datetime.datetime.strptime(date_string, '%d.%m.%y %H:%M:%S')
        elements = result[0].split(' ')
        steps = [elements[n:n + 7] for n in range(0, len(elements), 7)]
        self.steps_with_results = []
        for step in steps:
            if len(step) > 1:
                del step[0]
                split_step_info = step[5].split('_')
                parsed_step = {
                    'name': split_step_info[2].replace('*', ' '),
                    'method': step[0],
                    'test_condition': float(step[1]),
                    'limit_value': float(step[2]),
                    'actual_condition': float(step[3]),
                    'actual_value': float(step[4]),
                    'test_duration': float(split_step_info[1]),
                }
                parsed_step['go'] = 'GO' if parsed_step['actual_value'] <= parsed_step['limit_value'] else 'NGO'
                self.steps_with_results.append(parsed_step)

    def store_as_xlsx(self, name):
        wb = Workbook()

        bold_font = Font(bold=True)

        dest_filename = name if name.endswith('.xlsx') else f'{name}.xlsx'

        ws1 = wb.active
        ws1.title = "Test Report"

        ws1['A1'] = 'Preset Name'
        ws1['B1'] = 'Date'
        ws1['A1'].font = bold_font
        ws1['B1'].font = bold_font

        ws1['A2'] = self.name
        ws1['B2'] = self.date

        ws1['A4'] = 'Step Number'
        ws1['B4'] = 'Method'
        ws1['C4'] = 'Step Name'
        ws1['D4'] = 'Limit Value'
        ws1['E4'] = 'Actual Value'
        ws1['F4'] = 'Test Condition'
        ws1['G4'] = 'Actual Condition'
        ws1['H4'] = 'Test Duration'
        ws1['I4'] = 'Go'

        ws1['A4'].font = bold_font
        ws1['B4'].font = bold_font
        ws1['C4'].font = bold_font
        ws1['D4'].font = bold_font
        ws1['E4'].font = bold_font
        ws1['F4'].font = bold_font
        ws1['G4'].font = bold_font
        ws1['H4'].font = bold_font
        ws1['I4'].font = bold_font

        row = 5
        for step in self.steps_with_results:
            for col in range(1, len(step) + 2):
                value = 0
                if col == 1:
                    value = row - 4
                if col == 2:
                    value = step['method']
                elif col == 3:
                    value = step['name']
                elif col == 4:
                    value = str(step['limit_value']) + ' mA'
                elif col == 5:
                    value = str(step['actual_value']) + ' mA'
                elif col == 6:
                    value = str(step['test_condition']) + ' V'
                elif col == 7:
                    value = str(step['actual_condition']) + ' V'
                elif col == 8:
                    value = str(step['test_duration']) + ' s'
                elif col == 9:
                    value = step['go']
                _ = ws1.cell(column=col, row=row, value=value)
            row += 1

        column_widths = []
        for row in ws1.iter_rows():
            for i, cell in enumerate(row):
                try:
                    column_widths[i] = max(column_widths[i], len(as_text(cell.value)))
                except IndexError:
                    column_widths.append(len(as_text(cell.value)))

        for i, column_width in enumerate(column_widths):
            ws1.column_dimensions[get_column_letter(i + 1)].width = column_width + 5

        wb.save(filename=dest_filename)

    def __str__(self):
        string = ''
        string += self.name + ' | ' + str(self.date) + '\n'
        for idx, step in enumerate(self.steps_with_results):
            string += str(step) + '\n'
        return string


class TestingDevice(ABC):

    @abstractmethod
    def __init__(self, serial_port: str):
        pass

    @abstractmethod
    def reconnect(self):
        pass

    @abstractmethod
    def send_custom_command(self, command_hex):
        pass

    @abstractmethod
    def read_all(self):
        pass

    @abstractmethod
    def beep(self):
        pass

    @abstractmethod
    def identify(self):
        pass

    @abstractmethod
    def get_first_available_report(self):
        pass

    @abstractmethod
    def get_all_reports(self):
        pass

    @abstractmethod
    def is_testing(self):
        pass

    @abstractmethod
    def start_test(self):
        pass

    @abstractmethod
    def clear_all_reports(self):
        pass

    @abstractmethod
    def close_communication(self):
        pass


class FakeTestingDevice(TestingDevice):

    def __init__(self, serial_port: str):
        self.port = serial_port
        self.id_string = "DEBUG DEVICE"

    def reconnect(self):
        pass

    def send_custom_command(self, command_hex):
        pass

    def read_all(self):
        pass

    def beep(self):
        pass

    def identify(self):
        pass

    def get_first_available_report(self):
        pass

    def get_all_reports(self):
        pass

    def is_testing(self):
        pass

    def start_test(self):
        pass

    def clear_all_reports(self):
        pass

    def close_communication(self):
        pass


class ActualTestingDevice(TestingDevice):

    BEEP_COMMAND = [0x02, 0x81, 0xfa, 0x62, 0x20, 0x33, 0x42, 0x03]
    IDENTIFY_COMMAND = [0x02, 0x81, 0xfd]
    GET_REPORT_COMMAND = [0x02, 0x81, 0x06]
    START_TEST_COMMAND = [0x02, 0x81, 0xfa, 0x73, 0x20, 0x32, 0x41, 0x03]

    def __init__(self, serial_port: str):
        self.port = serial_port
        self.ser = serial.Serial(serial_port, baudrate=9600, timeout=None, parity=serial.PARITY_NONE,
                                 bytesize=serial.EIGHTBITS, stopbits=serial.STOPBITS_ONE, xonxoff=False)
        self.id_string = ""

    def reconnect(self):
        logging.debug('Trying to reconnect...')
        self.ser.close()
        try:
            self.ser.open()
            logging.info('Succesfully reconnected on {0}'.format(self.port))
        except serial.SerialException:
            i = 0
            base_serial_string = "/dev/ttyUSB"
            while i < 100:
                try:
                    i += 1
                    device = ActualTestingDevice(base_serial_string + str(i))
                    id_string = device.identify()
                    # this may need to be improved by actually checking the response
                    if id_string == self.id_string:
                        break
                except serial.SerialException:
                    continue
                except Exception:
                    logging.exception('Unknown exception while searching for testing device.')
            if i < 100:
                self.port = base_serial_string + str(i)
                self.ser = serial.Serial(self.port, baudrate=9600, timeout=None, parity=serial.PARITY_NONE,
                                    bytesize=serial.EIGHTBITS, stopbits=serial.STOPBITS_ONE, xonxoff=False)
                logging.info('Succesfully reconnected on {0}'.format(self.port))

    def send_custom_command(self, command_hex):
        try:
            self.ser.write(serial.to_bytes(command_hex))
        except serial.SerialException or OSError as e:
            logging.exception('Exception while writing command. Maybe the device was disconnected?')
            raise e
        time.sleep(0.12)

    def read_all(self):
        read_data = ''
        try:
            while self.ser.in_waiting > 0:
                try:
                    new_data = self.ser.read(self.ser.in_waiting)
                    read_data += new_data.decode(errors='ignore')
                except serial.SerialException or OSError as e:
                    logging.exception('Exception while reading from serial device. Maybe it was disconnected?')
                    raise e
                except TypeError:
                    continue
                time.sleep(0.12)
        except serial.SerialException or OSError as e:
            logging.exception('Exception while reading from serial device. Maybe it was disconnected?')
            raise e
        if len(read_data.strip()) > 0:
            logging.debug('Read {0} bytes from device: {1}'.format(len(read_data), ":".join("{:02x}".format(ord(c)) for c in read_data)))
        return read_data

    def beep(self):
        self.send_custom_command(ActualTestingDevice.BEEP_COMMAND)

    def identify(self):
        logging.debug('Requesting identification.')
        self.send_custom_command(ActualTestingDevice.IDENTIFY_COMMAND)
        id_string = self.read_all()
        id_string = (id_string.split("Conness.")[0])[3:]
        self.id_string = id_string
        logging.debug('Device identifies as {0}'.format(self.id_string))
        return id_string

    def get_first_available_report(self):
        self.send_custom_command(ActualTestingDevice.GET_REPORT_COMMAND)
        result = self.read_all()
        if len(result.strip().replace('\x07', '').replace('\x15', '').replace('4', '').replace('\x03', '').replace('2', '')) == 0:
            raise NoReportException('No report available for download.')
        return TestReport(result)

    def get_all_reports(self):
        reports = []
        while True:
            try:
                reports.append(self.get_first_available_report())
            except NoReportException:
                return reports

    def is_testing(self):
        self.beep()
        time.sleep(0.12)
        result = self.read_all()
        # some magic conversion
        return (":".join("{:02x}".format(ord(c)) for c in result)) == "07"

    def start_test(self):
        self.send_custom_command(ActualTestingDevice.START_TEST_COMMAND)
        self.ser.reset_input_buffer()

    def clear_all_reports(self):
        self.get_all_reports()

    def close_communication(self):
        self.ser.close()


class TextFeedback(QtCore.QObject):

    text_feedback_update = QtCore.pyqtSignal(str)

    def __init__(self):
        super().__init__()
        self.text = ''

    def set_text(self, text: str):
        self.text = text
        self.text_feedback_update.emit(text)

    def clear(self):
        self.set_text('')

    def append(self, text: str):
        self.set_text(self.text + text)
        pass

    def append_new_line(self, text: str):
        if len(self.text) == 0:
            self.append(text)
        else:
            self.append('\n' + text)


class StatusFeedback(QtCore.QObject):

    status_feedback_update = QtCore.pyqtSignal(str)

    def __init__(self):
        super().__init__()
        self.text = ''

    def set_text(self, text: str):
        self.text = text
        self.status_feedback_update.emit(text)


class LoadingIndicator(QtCore.QObject):

    set_loading_indicator_enable = QtCore.pyqtSignal(bool)

    def __init__(self):
        super().__init__()
        self.enabled = True

    def enable(self):
        self.enabled = True
        self.set_loading_indicator_enable.emit(True)

    def disable(self):
        self.enabled = False
        self.set_loading_indicator_enable.emit(False)

    def toggle_enable(self):
        if self.enabled:
            self.disable()
        else:
            self.enable()


class StartTestControl(QtCore.QObject):

    set_start_test_enable = QtCore.pyqtSignal(bool)

    def __init__(self):
        super().__init__()
        self.enabled = True

    def enable(self):
        self.enabled = True
        self.set_start_test_enable.emit(True)

    def disable(self):
        self.enabled = False
        self.set_start_test_enable.emit(False)

    def toggle_enable(self):
        if self.enabled:
            self.disable()
        else:
            self.enable()


class TestManager(QtCore.QThread):

    show_filename_dialog = QtCore.pyqtSignal(int)
    unexpected_shutdown_detected = QtCore.pyqtSignal(int)
    communication_error = QtCore.pyqtSignal(int)

    POLLING_INTERVAL = 5
    TEMP_FOLDER = 'temp'

    def on_reconnect_signal(self, number: int):
        self.device.reconnect()

    def on_startup(self, number: int):
        self.status_feedback.set_text("Connected to {0}".format(self.device.identify()))
        if os.path.exists(self.TEMP_FOLDER + '/test_running'):
            logging.warning('Unexpected shutdown detected.')
            self.unexpected_shutdown_detected.emit(1)

    def on_filename_available(self, filename: str):
        if filename:
            self.last_report.store_as_xlsx(filename)
            self.text_feedback.append_new_line("Report was saved to {0}".format(filename))
        else:
            self.text_feedback.append_new_line("Report was NOT saved. Please note that a backup copy was stored"
                                               " in {0}/backups".format(os.getcwd()))
        self.end_test()

    def resume(self):
        logging.info('Resuming...')
        self.start_test_control.disable()
        self.loading_indicator.enable()
        self.text_feedback.append_new_line("Unexpected shutdown detected. Waiting for report...")
        self.wait_for_report()

    def on_should_resume(self, should_resume: bool):
        if should_resume:
            self.please_resume = True
        else:
            os.remove(self.TEMP_FOLDER + '/test_running')

    def __init__(self, device: ActualTestingDevice, config):
        super().__init__()

        logging.basicConfig(**config.log_config)

        if not os.path.exists(self.TEMP_FOLDER):
            os.makedirs(self.TEMP_FOLDER)

        self.backup_folder = config.backup_folder

        if not os.path.exists(self.backup_folder):
            os.makedirs(self.backup_folder)

        self.backup_folder_max_size = config.backup_folder_max_size

        self.please_resume = False

        self.device = device
        self.text_feedback: TextFeedback = TextFeedback()
        self.status_feedback: StatusFeedback = StatusFeedback()
        self.start_test_control: StartTestControl = StartTestControl()
        self.loading_indicator: LoadingIndicator = LoadingIndicator()
        self.last_report = None

    def clean_backup_folder(self):
        size = sum(os.path.getsize(self.backup_folder + '/' + f) for f in os.listdir(self.backup_folder) if os.path.isfile(self.backup_folder + '/' + f))
        logging.debug('Backup folder size is {0}'.format(size))
        if size > self.backup_folder_max_size:
            logging.info('Backup folder max size exceeded. Purging backups starting from the oldest.')
            files = os.listdir(self.backup_folder)
            files.sort()
            i = 0
            while size > self.backup_folder_max_size:
                file_size = os.path.getsize(self.backup_folder + '/' + files[i])
                os.remove(self.backup_folder + '/' + files[i])
                size -= file_size
                i += 1
            logging.info('{0} files deleted.'.format(i))

    def end_test(self):
        self.start_test_control.enable()
        self.loading_indicator.disable()
        os.remove(self.TEMP_FOLDER + '/test_running')
        self.please_resume = False

    def wait_for_report(self):
        self.clean_backup_folder()
        try:
            while self.device.is_testing():
                time.sleep(self.POLLING_INTERVAL)

            report = self.device.get_first_available_report()
            self.text_feedback.append_new_line("Report downloaded succesfully.")
            self.last_report = report
            report.store_as_xlsx("{0}/{1}.xlsx".format(self.backup_folder, (int(time.time() * 1000))))
            self.show_filename_dialog.emit(1)
        except serial.SerialException or OSError:
            self.communication_error.emit(1)
        except NoReportException:
            self.text_feedback.append_new_line('Test stopped. Ready for new test.')
            self.end_test()

    def start_test(self):
        self.device.get_all_reports()
        try:
            self.device.start_test()
            self.start_test_control.disable()
            self.loading_indicator.enable()
            Path(self.TEMP_FOLDER + '/test_running').touch()
            self.text_feedback.clear()
            self.text_feedback.append_new_line("Test started.")
            self.text_feedback.append_new_line("Waiting for report...")
            self.wait_for_report()
        except serial.SerialException:
            self.communication_error.emit(1)

    def run(self):
        if self.please_resume:
            self.resume()
        else:
            self.start_test()
